import bs4
import openpyxl

from itertools import chain


# 抓取 PTT 電影版的網頁原始碼（HTML）
import urllib.request as req

from openpyxl import Workbook
url = "https://tw.stock.yahoo.com/rank/price"
# 建立一個 Request 物件, 附加 Request Headers 的資訊
request = req.Request(url, headers={
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36"
})
with req.urlopen(request) as response:
    data = response.read().decode("utf-8")

# 解析原始碼, 取得每篇文章的標題
root = bs4.BeautifulSoup(data, "html.parser")
# 尋找所有 class = "title" 的 div 標籤
brands = root.find_all("div", class_="Lh(20px) Fw(600) Fz(16px) Ell")
brandNumbers = root.find_all("div", class_="D(f) Ai(c)")
closingPrices = root.find_all("div", class_="")


# for brand in brands :
#	if brand.string != None : # 如果標題不是空值, 印出來
#			print(brand.string)

# for brandNumber in brandNumbers:
#	if brandNumber.string != None :
#		print(brandNumber.string)

for meow in chain(range(len(brands)), range(len(brandNumbers))):
    print(brands[meow].string+"  "+brandNumbers[meow].string)


searchTitle = input('你想搜尋哪支股票:')
for meow in chain(range(len(brands))):
    if(brands[meow].string.find(searchTitle) != -1):
        print(brands[meow].string+"  "+brandNumbers[meow].string)


# 將爬蟲資料存成 Excel 檔

# 建立新活頁簿
wb = Workbook()

# 顯示工作表名稱
# print(wb.sheetnames)
sheet = wb['Sheet']


titles = ("公司名稱", "公司代碼")
sheet.append(titles)

for i in range(len(brands)):
    sheet.cell(row=i+1, column=1, value=brands[i].string)

for j in range(len(brandNumbers)):
    sheet.cell(row=j+1, column=2, value=brandNumbers[j].string)


sheet.insert_rows(1, 1)
sheet.cell(row=1, column=1, value="公司名稱")
sheet.cell(row=1, column=2, value="公司代號")
wb.save("yahoostock.xlsx")


# 執行： python crawlerHW1.py
